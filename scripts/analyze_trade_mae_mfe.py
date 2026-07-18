#!/usr/bin/env python3
"""MAE/MFE trade excursion report from Excel trade log (read-only research).

Reads ``Trade Log`` sheet, fetches Upstox 5m→10m candles, computes MAE/MFE,
round-trip / post-exit reversion / pre-entry flags. Writes HTML (rscd-style)
+ JSON under docs/diagnostics. Does NOT touch live gates or consistency logs.

Run (prefer paperclip app container — needs Upstox token):
  PYTHONPATH=. python3 scripts/analyze_trade_mae_mfe.py
  PYTHONPATH=. python3 scripts/analyze_trade_mae_mfe.py \\
      --xlsx docs/diagnostics/trade_log_13Jul_17Jul2026.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pytz

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

IST = pytz.timezone("Asia/Kolkata")
PRE_BARS = 5
POST_BARS = 5
ROUND_TRIP_MFE_MULT = 1.5
REVERSION_MULT = 1.0
# Near-zero SL gaps make R-multiples unstable (entry ≈ VWAP). Still compute,
# but flag and exclude from average MAE/MFE so the week summary isn't skewed.
MIN_STABLE_RISK_PTS = 0.5
MIN_STABLE_RISK_PCT = 0.05  # percent of entry price

DEFAULT_XLSX = _ROOT / "docs/diagnostics/trade_log_13Jul_17Jul2026.xlsx"
DEFAULT_OUT_DIR = _ROOT / "docs/diagnostics"


@dataclass
class TradeRow:
    idx: int
    session_date: date
    symbol: str
    direction: str
    entry_time: time
    entry_price: float
    exit_time: time
    exit_price: float
    ema10_sl: Optional[float]
    vwap_sl: Optional[float]
    grade: str
    notes: str


@dataclass
class TradeAnalysis:
    trade: Dict[str, Any]
    skip_reason: Optional[str] = None
    instrument_key: Optional[str] = None
    risk_ema10_pts: Optional[float] = None
    risk_vwap_pts: Optional[float] = None
    effective_risk_pts: Optional[float] = None
    effective_sl_ref: Optional[str] = None
    effective_sl_price: Optional[float] = None
    sl_choice_reason: Optional[str] = None
    mae_pts: Optional[float] = None
    mae_mult: Optional[float] = None
    mae_on_exit_candle: Optional[bool] = None
    mae_recovered: Optional[bool] = None
    mfe_pts: Optional[float] = None
    mfe_mult: Optional[float] = None
    mfe_bars_after_entry: Optional[int] = None
    round_trip: bool = False
    round_trip_peak_mult: Optional[float] = None
    post_exit_reversion: bool = False
    post_exit_reversion_mult: Optional[float] = None
    exit_candle_range_pts: Optional[float] = None
    exit_candle_range_mult: Optional[float] = None
    pnl_pts: Optional[float] = None
    pnl_r: Optional[float] = None
    atr14_pts: Optional[float] = None
    mae_atr: Optional[float] = None
    mfe_atr: Optional[float] = None
    exit_candle_range_atr: Optional[float] = None
    single_candle_blowout_atr: bool = False
    pre_entry: Dict[str, Any] = field(default_factory=dict)
    flags: List[str] = field(default_factory=list)


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(v: Any) -> Optional[time]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.time().replace(second=0, microsecond=0)
    if isinstance(v, time):
        return v.replace(second=0, microsecond=0)
    s = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def load_trades(xlsx: Path) -> List[TradeRow]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "Trade Log" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Trade Log' not found; have {wb.sheetnames}")
    ws = wb["Trade Log"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    want = {
        "Date",
        "Symbol",
        "Direction",
        "Entry Time",
        "Entry Price",
        "Exit Time",
        "Exit Price",
        "Planned SL (EMA10 at entry)",
        "VWAP (SL if EMA10 is far)",
        "Confidence Grade",
        "Notes",
    }
    if set(headers) < want:
        missing = want - set(headers)
        raise SystemExit(f"Missing columns: {sorted(missing)}")
    col = {h: i for i, h in enumerate(headers)}
    out: List[TradeRow] = []
    for i, r in enumerate(rows[1:], start=2):
        if not r or not any(c is not None and str(c).strip() for c in r):
            continue
        d = _parse_date(r[col["Date"]])
        sym = str(r[col["Symbol"]] or "").strip().upper()
        direction = str(r[col["Direction"]] or "").strip().upper()
        et = _parse_time(r[col["Entry Time"]])
        xt = _parse_time(r[col["Exit Time"]])
        ep = _f(r[col["Entry Price"]])
        xp = _f(r[col["Exit Price"]])
        if not d or not sym or direction not in ("LONG", "SHORT") or et is None or xt is None:
            continue
        if ep is None or xp is None:
            continue
        out.append(
            TradeRow(
                idx=i,
                session_date=d,
                symbol=sym,
                direction=direction,
                entry_time=et,
                entry_price=float(ep),
                exit_time=xt,
                exit_price=float(xp),
                ema10_sl=_f(r[col["Planned SL (EMA10 at entry)"]]),
                vwap_sl=_f(r[col["VWAP (SL if EMA10 is far)"]]),
                grade=str(r[col["Confidence Grade"]] or "").strip() or "—",
                notes=str(r[col["Notes"]] or "").strip(),
            )
        )
    return out


def resolve_instrument_key(db, symbol: str) -> Optional[str]:
    from sqlalchemy import text

    row = db.execute(
        text(
            """
            SELECT currmth_future_instrument_key AS ikey
            FROM arbitrage_master
            WHERE UPPER(stock) = :s
            LIMIT 1
            """
        ),
        {"s": symbol.upper()},
    ).fetchone()
    return str(row.ikey) if row and row.ikey else None


def fetch_5m_candles(ikey: str, session_date: date) -> List[Dict[str, Any]]:
    from backend.config import settings
    from backend.services.relative_strength_scanner import CANDLE_INTERVAL, _sorted_candles
    from backend.services.upstox_service import UpstoxService

    ux = UpstoxService(settings.UPSTOX_API_KEY, settings.UPSTOX_API_SECRET)
    raw = ux.get_historical_candles_by_instrument_key(
        ikey,
        interval=CANDLE_INTERVAL,
        days_back=8,
        range_end_date=session_date,
    )
    if not raw:
        return []
    return _sorted_candles(raw)


def _bar_end_ist(bar: Dict[str, Any]) -> Optional[datetime]:
    from backend.services.kavach_volume import _parse_ist

    be = bar.get("bar_end")
    if isinstance(be, datetime):
        return be.astimezone(IST) if be.tzinfo else IST.localize(be)
    ts = _parse_ist(bar.get("timestamp"))
    if ts is None:
        return None
    return ts + timedelta(minutes=5)


def atr14_wilder(
    highs: List[float], lows: List[float], closes: List[float], period: int = 14
) -> List[Optional[float]]:
    """Wilder ATR(period). Early bars use expanding mean so entries always have a denom."""
    n = len(closes)
    if n == 0:
        return []
    tr: List[float] = []
    for i in range(n):
        if i == 0:
            tr.append(max(0.0, highs[i] - lows[i]))
        else:
            tr.append(
                max(
                    highs[i] - lows[i],
                    abs(highs[i] - closes[i - 1]),
                    abs(lows[i] - closes[i - 1]),
                )
            )
    out: List[Optional[float]] = [None] * n
    if n < period:
        s = 0.0
        for i in range(n):
            s += tr[i]
            out[i] = s / (i + 1)
        return out
    s = 0.0
    for i in range(period - 1):
        s += tr[i]
        out[i] = s / (i + 1)
    prev = sum(tr[:period]) / period
    out[period - 1] = prev
    for i in range(period, n):
        prev = (prev * (period - 1) + tr[i]) / period
        out[i] = prev
    return out


def build_session_10m(
    candles_5m: List[Dict[str, Any]], session_date: date
) -> List[Dict[str, Any]]:
    from backend.services.kavach_10m import aggregate_10m_bars
    from backend.services.relative_strength_scanner import _parse_ist_date
    from backend.services.vajra.indicators import cumulative_vwap, ema_series

    sess = session_date.isoformat()
    all_10m = aggregate_10m_bars(candles_5m)
    if not all_10m:
        return []

    # ATR from multi-day 10m path so morning entries aren't starved of history.
    atr_all = atr14_wilder(
        [float(b["high"]) for b in all_10m],
        [float(b["low"]) for b in all_10m],
        [float(b["close"]) for b in all_10m],
        14,
    )
    atr_by_ts = {
        str(b.get("timestamp")): atr_all[i]
        for i, b in enumerate(all_10m)
        if i < len(atr_all)
    }

    bars = [b for b in all_10m if _parse_ist_date(b.get("timestamp")) == sess]
    if not bars:
        return []

    closes = [float(b["close"]) for b in bars]
    ema5_s = ema_series(closes, 5)
    ema10_s = ema_series(closes, 10)

    for i, b in enumerate(bars):
        g_end = int(b["end_5m_idx"])
        slice_5 = [
            c
            for j, c in enumerate(candles_5m)
            if j <= g_end and _parse_ist_date(c.get("timestamp")) == sess
        ]
        vwap = None
        if slice_5:
            hs = [_f(c.get("high")) or 0.0 for c in slice_5]
            ls = [_f(c.get("low")) or 0.0 for c in slice_5]
            cs = [_f(c.get("close")) or 0.0 for c in slice_5]
            vs = [_f(c.get("volume")) or 0.0 for c in slice_5]
            vwap_s = cumulative_vwap(hs, ls, cs, vs)
            vwap = vwap_s[-1] if vwap_s else None
        b["ema5"] = ema5_s[i] if i < len(ema5_s) else None
        b["ema10"] = ema10_s[i] if i < len(ema10_s) else None
        b["vwap"] = vwap
        b["atr14"] = atr_by_ts.get(str(b.get("timestamp")))
        b["bar_end_ist"] = _bar_end_ist(b)
        b["i"] = i
    return bars


def bar_index_at_or_after(bars: List[Dict[str, Any]], when: datetime) -> Optional[int]:
    if when.tzinfo is None:
        when = IST.localize(when)
    else:
        when = when.astimezone(IST)
    for i, b in enumerate(bars):
        be = b.get("bar_end_ist")
        if be is not None and be >= when:
            return i
    return None


def bar_index_containing(bars: List[Dict[str, Any]], when: datetime) -> Optional[int]:
    if when.tzinfo is None:
        when = IST.localize(when)
    else:
        when = when.astimezone(IST)
    for i, b in enumerate(bars):
        be = b.get("bar_end_ist")
        if be is None:
            continue
        start = be - timedelta(minutes=10)
        if start < when <= be:
            return i
    return bar_index_at_or_after(bars, when)


def notes_force_sl_ref(notes: str) -> Optional[str]:
    n = (notes or "").lower()
    if re.search(r"vwap\s+as\s+sl|considered\s+vwap\s+as\s+sl|sl\s*[:=]\s*vwap", n):
        return "VWAP"
    if re.search(r"ema10\s+as\s+sl|sl\s*[:=]\s*ema10|planned\s+sl\s*[:=]\s*ema", n):
        return "EMA10"
    return None


def choose_effective_sl(
    tr: TradeRow,
) -> Tuple[Optional[float], Optional[str], Optional[float], Optional[float], str]:
    entry = tr.entry_price
    r_ema = abs(entry - tr.ema10_sl) if tr.ema10_sl is not None else None
    r_vwap = abs(entry - tr.vwap_sl) if tr.vwap_sl is not None else None
    forced = notes_force_sl_ref(tr.notes)
    if forced == "VWAP" and r_vwap is not None:
        return r_vwap, "VWAP", r_ema, r_vwap, "Notes specify VWAP as SL"
    if forced == "EMA10" and r_ema is not None:
        return r_ema, "EMA10", r_ema, r_vwap, "Notes specify EMA10 as SL"
    if r_ema is None and r_vwap is None:
        return None, None, None, None, "Both SL references missing"
    if r_ema is None:
        return r_vwap, "VWAP", r_ema, r_vwap, "EMA10 missing — VWAP only"
    if r_vwap is None:
        return r_ema, "EMA10", r_ema, r_vwap, "VWAP missing — EMA10 only"
    if r_vwap < r_ema - 1e-9:
        return r_vwap, "VWAP", r_ema, r_vwap, "VWAP risk smaller (default nearer stop)"
    if r_ema < r_vwap - 1e-9:
        return r_ema, "EMA10", r_ema, r_vwap, "EMA10 risk smaller (default nearer stop)"
    return r_ema, "EMA10", r_ema, r_vwap, "EMA10 and VWAP risk equal — prefer EMA10"


def adverse_pts(is_long: bool, entry: float, low: float, high: float) -> float:
    if is_long:
        return max(0.0, entry - low)
    return max(0.0, high - entry)


def favor_pts(is_long: bool, entry: float, low: float, high: float) -> float:
    if is_long:
        return max(0.0, high - entry)
    return max(0.0, entry - low)


# Stable codes for pre-entry sub-flags (used in win/loss discrimination).
PRE_VOL_DECEL = "volume_decel"
PRE_BODY_SHRINK = "body_shrink"
PRE_EMA_VWAP_MISALIGN = "ema_vwap_misalign"
PRE_WIDE_GAP = "wide_ema_vwap_gap"
PRE_EMA5_SEP_FADE = "ema5_sep_fade"
PRE_CODE_LABELS = {
    PRE_VOL_DECEL: "volume decelerating into entry",
    PRE_BODY_SHRINK: "candle body shrinking into entry (momentum fade)",
    PRE_EMA_VWAP_MISALIGN: "EMA10/VWAP misaligned with trade direction",
    PRE_WIDE_GAP: "wide EMA10–VWAP gap pre-entry",
    PRE_EMA5_SEP_FADE: "price losing separation from EMA5 into entry",
}


def pre_entry_check(pre_bars: List[Dict[str, Any]], is_long: bool) -> Dict[str, Any]:
    if not pre_bars:
        return {
            "warning": False,
            "summary": "no pre-entry bars",
            "details": [],
            "warnings": [],
            "warning_codes": [],
        }

    details = []
    codes: List[str] = []
    for b in pre_bars:
        ema10 = b.get("ema10")
        vwap = b.get("vwap")
        gap = None
        side = None
        if ema10 is not None and vwap is not None:
            gap = abs(float(ema10) - float(vwap))
            if float(ema10) > float(vwap):
                side = "EMA10>VWAP"
            elif float(ema10) < float(vwap):
                side = "EMA10<VWAP"
            else:
                side = "EMA10=VWAP"
        details.append(
            {
                "bar_end": str(b.get("bar_end_ist") or ""),
                "close": round(float(b["close"]), 2),
                "volume": round(float(b.get("volume") or 0), 0),
                "range": round(float(b["high"]) - float(b["low"]), 2),
                "ema10": round(float(ema10), 2) if ema10 is not None else None,
                "vwap": round(float(vwap), 2) if vwap is not None else None,
                "ema10_vwap": side,
                "ema10_vwap_gap": round(gap, 2) if gap is not None else None,
            }
        )

    if len(pre_bars) >= 4:
        early = pre_bars[:-2]
        late = pre_bars[-2:]

        def _body(b: Dict[str, Any]) -> float:
            return abs(float(b["close"]) - float(b["open"]))

        early_vol = sum(float(b.get("volume") or 0) for b in early) / len(early)
        late_vol = sum(float(b.get("volume") or 0) for b in late) / len(late)
        early_body = sum(_body(b) for b in early) / len(early)
        late_body = sum(_body(b) for b in late) / len(late)
        if early_vol > 0 and late_vol < 0.65 * early_vol:
            codes.append(PRE_VOL_DECEL)
        if early_body > 0 and late_body < 0.65 * early_body:
            codes.append(PRE_BODY_SHRINK)

    last = pre_bars[-1]
    ema10 = last.get("ema10")
    vwap = last.get("vwap")
    gap_pct = None
    if ema10 is not None and vwap is not None:
        if (is_long and float(ema10) < float(vwap)) or (
            (not is_long) and float(ema10) > float(vwap)
        ):
            codes.append(PRE_EMA_VWAP_MISALIGN)
        gap_pct = abs(float(ema10) - float(vwap)) / max(abs(float(last["close"])), 1e-9) * 100
        if gap_pct > 0.35:
            codes.append(PRE_WIDE_GAP)

    if len(pre_bars) >= 3 and all(b.get("ema5") is not None for b in pre_bars[-3:]):
        dists = []
        for b in pre_bars[-3:]:
            if is_long:
                dists.append(float(b["close"]) - float(b["ema5"]))
            else:
                dists.append(float(b["ema5"]) - float(b["close"]))
        if dists[0] > 0 and dists[-1] < dists[0] * 0.4:
            codes.append(PRE_EMA5_SEP_FADE)

    warnings = []
    for c in codes:
        label = PRE_CODE_LABELS.get(c, c)
        if c == PRE_WIDE_GAP and gap_pct is not None:
            warnings.append(f"{label} ({gap_pct:.2f}% of price)")
        else:
            warnings.append(label)

    summary = "; ".join(warnings) if warnings else "no clear pre-entry red flag from 10m structure"
    return {
        "warning": bool(codes),
        "summary": summary,
        "details": details,
        "warnings": warnings,
        "warning_codes": codes,
    }


def _trade_dict(tr: TradeRow) -> Dict[str, Any]:
    return {
        "row": tr.idx,
        "date": tr.session_date.isoformat(),
        "symbol": tr.symbol,
        "direction": tr.direction,
        "entry_time": tr.entry_time.strftime("%H:%M"),
        "exit_time": tr.exit_time.strftime("%H:%M"),
        "entry_price": tr.entry_price,
        "exit_price": tr.exit_price,
        "grade": tr.grade,
        "notes": tr.notes,
        "ema10_sl": tr.ema10_sl,
        "vwap_sl": tr.vwap_sl,
    }


def analyze_one(tr: TradeRow, bars: List[Dict[str, Any]], ikey: str) -> TradeAnalysis:
    is_long = tr.direction == "LONG"
    entry_dt = IST.localize(datetime.combine(tr.session_date, tr.entry_time))
    exit_dt = IST.localize(datetime.combine(tr.session_date, tr.exit_time))
    if exit_dt < entry_dt:
        exit_dt += timedelta(days=1)

    risk, ref, r_ema, r_vwap, reason = choose_effective_sl(tr)
    sl_price = tr.vwap_sl if ref == "VWAP" else tr.ema10_sl
    base = TradeAnalysis(
        trade=_trade_dict(tr),
        instrument_key=ikey,
        risk_ema10_pts=round(r_ema, 4) if r_ema is not None else None,
        risk_vwap_pts=round(r_vwap, 4) if r_vwap is not None else None,
        effective_risk_pts=round(risk, 4) if risk is not None else None,
        effective_sl_ref=ref,
        effective_sl_price=sl_price,
        sl_choice_reason=reason,
    )

    if not bars:
        base.skip_reason = "no 10m bars for session"
        return base
    if risk is None or risk <= 0:
        base.skip_reason = "effective planned risk unavailable/zero"
        return base

    min_stable = max(MIN_STABLE_RISK_PTS, abs(tr.entry_price) * MIN_STABLE_RISK_PCT / 100.0)
    if risk < min_stable:
        base.flags.append("DEGENERATE_RISK")
        base.sl_choice_reason = (
            f"{base.sl_choice_reason} — near-zero gap ({risk:.2f} pts); R multiples unstable"
        )

    i_entry = bar_index_containing(bars, entry_dt)
    i_exit = bar_index_containing(bars, exit_dt)
    if i_entry is None or i_exit is None:
        base.skip_reason = "could not map entry/exit to 10m bars"
        return base
    if i_exit < i_entry:
        i_exit = i_entry

    trade_bars = bars[i_entry : i_exit + 1]
    pre = bars[max(0, i_entry - PRE_BARS) : i_entry]
    post = bars[i_exit + 1 : i_exit + 1 + POST_BARS]

    mae = 0.0
    mfe = 0.0
    mfe_bar_offset = 0
    mae_bar_i = i_entry
    for off, b in enumerate(trade_bars):
        a = adverse_pts(is_long, tr.entry_price, float(b["low"]), float(b["high"]))
        fav = favor_pts(is_long, tr.entry_price, float(b["low"]), float(b["high"]))
        if a > mae:
            mae = a
            mae_bar_i = i_entry + off
        if fav > mfe:
            mfe = fav
            mfe_bar_offset = off

    base.mae_pts = round(mae, 4)
    base.mfe_pts = round(mfe, 4)
    base.mae_mult = round(mae / risk, 3)
    base.mfe_mult = round(mfe / risk, 3)
    base.mfe_bars_after_entry = mfe_bar_offset
    base.mae_on_exit_candle = mae_bar_i == i_exit
    base.mae_recovered = mae_bar_i < i_exit

    if is_long:
        pnl = tr.exit_price - tr.entry_price
    else:
        pnl = tr.entry_price - tr.exit_price
    base.pnl_pts = round(pnl, 4)
    base.pnl_r = round(pnl / risk, 3)
    if mfe >= ROUND_TRIP_MFE_MULT * risk and pnl <= 0:
        base.round_trip = True
        base.round_trip_peak_mult = round(mfe / risk, 3)
        base.flags.append("ROUND_TRIP")

    if post:
        if is_long:
            best = max(float(b["high"]) for b in post) - tr.exit_price
        else:
            best = tr.exit_price - min(float(b["low"]) for b in post)
        best = max(0.0, best)
        base.post_exit_reversion_mult = round(best / risk, 3)
        if best >= REVERSION_MULT * risk:
            base.post_exit_reversion = True
            base.flags.append("POST_EXIT_REVERSION")
    else:
        base.post_exit_reversion_mult = 0.0

    xb = bars[i_exit]
    rng = float(xb["high"]) - float(xb["low"])
    base.exit_candle_range_pts = round(rng, 4)
    base.exit_candle_range_mult = round(rng / risk, 3)
    if base.exit_candle_range_mult is not None and base.exit_candle_range_mult >= 2.0:
        base.flags.append("EXIT_CANDLE_BLOWOUT")
    if base.mae_mult is not None and base.mae_mult >= 2.0 and base.mae_on_exit_candle:
        base.flags.append("SINGLE_CANDLE_MAE_BLOWOUT")

    atr = bars[i_entry].get("atr14")
    if atr is not None and float(atr) > 0:
        atr_f = float(atr)
        base.atr14_pts = round(atr_f, 4)
        base.mae_atr = round(mae / atr_f, 3)
        base.mfe_atr = round(mfe / atr_f, 3)
        base.exit_candle_range_atr = round(rng / atr_f, 3)
        if base.mae_on_exit_candle and mae >= 2.0 * atr_f:
            base.single_candle_blowout_atr = True
            base.flags.append("SINGLE_CANDLE_MAE_BLOWOUT_ATR")

    pre_info = pre_entry_check(pre, is_long)
    base.pre_entry = pre_info
    if pre_info.get("warning"):
        base.flags.append("PRE_ENTRY_WARNING")

    return base


def _avg(vals: List[float]) -> Optional[float]:
    return round(sum(vals) / len(vals), 3) if vals else None


def _median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    s = sorted(vals)
    mid = len(s) // 2
    if len(s) % 2:
        return round(s[mid], 3)
    return round((s[mid - 1] + s[mid]) / 2.0, 3)


def _sl_excursion_metrics(items: List[TradeAnalysis]) -> Dict[str, Any]:
    """Points + ATR + R metrics for one SL-reference group (stable risk preferred for R)."""
    stable = [x for x in items if "DEGENERATE_RISK" not in x.flags]
    return {
        "n": len(items),
        "n_stable_risk": len(stable),
        "avg_effective_risk_pts": _avg(
            [x.effective_risk_pts for x in items if x.effective_risk_pts is not None]
        ),
        "median_effective_risk_pts": _median(
            [x.effective_risk_pts for x in items if x.effective_risk_pts is not None]
        ),
        "avg_mae_pts": _avg([x.mae_pts for x in items if x.mae_pts is not None]),
        "median_mae_pts": _median([x.mae_pts for x in items if x.mae_pts is not None]),
        "avg_mfe_pts": _avg([x.mfe_pts for x in items if x.mfe_pts is not None]),
        "median_mfe_pts": _median([x.mfe_pts for x in items if x.mfe_pts is not None]),
        "avg_exit_range_pts": _avg(
            [x.exit_candle_range_pts for x in items if x.exit_candle_range_pts is not None]
        ),
        "avg_atr14_pts": _avg([x.atr14_pts for x in items if x.atr14_pts is not None]),
        "avg_mae_atr": _avg([x.mae_atr for x in items if x.mae_atr is not None]),
        "median_mae_atr": _median([x.mae_atr for x in items if x.mae_atr is not None]),
        "avg_mfe_atr": _avg([x.mfe_atr for x in items if x.mfe_atr is not None]),
        "median_mfe_atr": _median([x.mfe_atr for x in items if x.mfe_atr is not None]),
        "avg_mae_mult_stable": _avg(
            [x.mae_mult for x in stable if x.mae_mult is not None]
        ),
        "avg_mfe_mult_stable": _avg(
            [x.mfe_mult for x in stable if x.mfe_mult is not None]
        ),
        "single_candle_blowout_r": sum(
            1 for x in items if "SINGLE_CANDLE_MAE_BLOWOUT" in x.flags
        ),
        "single_candle_blowout_atr": sum(1 for x in items if x.single_candle_blowout_atr),
        "post_exit_reversion": sum(1 for x in items if x.post_exit_reversion),
        "round_trip": sum(1 for x in items if x.round_trip),
    }


def _ratio_gap(a: Optional[float], b: Optional[float]) -> Optional[float]:
    """Relative gap (a-b)/b when both positive; None otherwise."""
    if a is None or b is None or b == 0:
        return None
    return round((a - b) / abs(b), 3)


def validate_denominator_artifact(by_sl: Dict[str, List[TradeAnalysis]]) -> Dict[str, Any]:
    """Compare EMA10 vs VWAP groups in R, points, and ATR — flag denominator bias."""
    ema = _sl_excursion_metrics(by_sl.get("EMA10") or [])
    vwap = _sl_excursion_metrics(by_sl.get("VWAP") or [])

    r_mae_gap = _ratio_gap(vwap.get("avg_mae_mult_stable"), ema.get("avg_mae_mult_stable"))
    pts_mae_gap = _ratio_gap(vwap.get("avg_mae_pts"), ema.get("avg_mae_pts"))
    med_pts_gap = _ratio_gap(vwap.get("median_mae_pts"), ema.get("median_mae_pts"))
    atr_mae_gap = _ratio_gap(vwap.get("avg_mae_atr"), ema.get("avg_mae_atr"))
    risk_gap = _ratio_gap(vwap.get("avg_effective_risk_pts"), ema.get("avg_effective_risk_pts"))

    # "Worse" = higher adverse excursion / higher blowout rate.
    worse_on_r = (r_mae_gap is not None and r_mae_gap > 0.15) or (
        (vwap.get("single_candle_blowout_r") or 0) > (ema.get("single_candle_blowout_r") or 0)
    )
    worse_on_avg_pts = pts_mae_gap is not None and pts_mae_gap > 0.15
    worse_on_med_pts = med_pts_gap is not None and med_pts_gap > 0.15
    # Prefer median for points verdict (avoids one fat-tail trade dominating).
    worse_on_pts = worse_on_med_pts if med_pts_gap is not None else worse_on_avg_pts
    pts_outlier_driven = worse_on_avg_pts and not worse_on_med_pts
    worse_on_atr = atr_mae_gap is not None and atr_mae_gap > 0.15
    blowout_atr_vwap = vwap.get("single_candle_blowout_atr") or 0
    blowout_atr_ema = ema.get("single_candle_blowout_atr") or 0
    worse_blowout_atr = blowout_atr_vwap > blowout_atr_ema

    controls_worse = worse_on_pts or worse_on_atr or worse_blowout_atr

    if worse_on_r and not controls_worse:
        extra = ""
        if pts_outlier_driven:
            extra = (
                " Avg MAE points look higher for VWAP, but medians are similar — that avg gap is "
                "outlier-driven, not a typical-trade effect."
            )
        verdict = (
            "DENOMINATOR ARTIFACT — VWAP-effective looks worse in R-multiples largely because "
            "its planned-risk denominator is tighter by construction. In raw points (median) and "
            "vs ATR(14), the adverse-excursion gap shrinks or disappears."
            + extra
            + " Do not treat the R-split alone as evidence that VWAP is a worse stop reference. "
            "Directional only (n=15)."
        )
        label = "artifact"
    elif worse_on_r and controls_worse:
        bits = []
        if worse_on_pts:
            bits.append("median MAE points")
        elif pts_outlier_driven:
            bits.append("avg MAE points (medians similar — outlier-driven)")
        if worse_on_atr:
            bits.append("MAE/ATR")
        if worse_blowout_atr:
            bits.append("ATR-normalized single-candle blowouts")
        # If only ATR/blowouts persist and points median does not:
        if (worse_on_atr or worse_blowout_atr) and not worse_on_pts:
            atr_bits = [b for b in bits if "points" not in b]
            pts_note = (
                " Avg MAE points look higher for VWAP but medians are nearly identical "
                "(outlier-driven, not typical)."
                if pts_outlier_driven
                else ""
            )
            verdict = (
                "PARTLY PERSISTS AFTER CONTROLLING — median MAE in points is similar across SL "
                "groups (R-gap is partly a denominator artifact), but VWAP-effective remains worse on "
                + (" and ".join(atr_bits) if atr_bits else "ATR-normalized measures")
                + "."
                + pts_note
                + " Treat as directional only (n=15), not a settled finding."
            )
            label = "persists"
        else:
            verdict = (
                "STILL LOOKS WORSE AFTER CONTROLLING — VWAP-effective remains worse on "
                + " and ".join(bits)
                + " (not only on R-multiples). The tighter denominator inflates R, but does not fully "
                "explain the split. Directional only (n=15)."
            )
            label = "persists"
    elif not worse_on_r:
        verdict = (
            "NO CLEAR R-SPLIT TO EXPLAIN — VWAP-effective is not materially worse on stable "
            "R-MAE in this sample; denominator-bias check is moot."
        )
        label = "no_r_split"
    else:
        verdict = (
            "MIXED — R-gap and points/ATR controls do not line up cleanly. Treat as inconclusive "
            "on this 15-trade sample."
        )
        label = "mixed"

    return {
        "sample_caveat": (
            "n=15 trades (13–17 Jul). All conclusions are directional only — not statistically reliable."
        ),
        "verdict_label": label,
        "verdict": verdict,
        "ema10": ema,
        "vwap": vwap,
        "gaps": {
            "vwap_vs_ema_avg_risk_pts": risk_gap,
            "vwap_vs_ema_avg_mae_r_stable": r_mae_gap,
            "vwap_vs_ema_avg_mae_pts": pts_mae_gap,
            "vwap_vs_ema_median_mae_pts": med_pts_gap,
            "vwap_vs_ema_avg_mae_atr": atr_mae_gap,
            "avg_mae_pts_outlier_driven": pts_outlier_driven,
        },
    }


def _flag_rate_block(
    winners: List[TradeAnalysis],
    losers: List[TradeAnalysis],
    predicate,
) -> Dict[str, Any]:
    def _rate(items: List[TradeAnalysis]) -> Optional[float]:
        if not items:
            return None
        return round(100.0 * sum(1 for x in items if predicate(x)) / len(items), 1)

    rw, rl = _rate(winners), _rate(losers)
    delta = None if rw is None or rl is None else round(rw - rl, 1)
    discriminating = None
    note = ""
    if delta is not None:
        if abs(delta) <= 15:
            discriminating = False
            note = (
                f"Flag rate on winners ({rw}%) vs losers ({rl}%) differs by {abs(delta):.0f} pp "
                f"(≤15 pp) — not currently discriminating."
            )
        else:
            discriminating = True
            side = "losers" if delta < 0 else "winners"
            note = (
                f"Flag rate on winners ({rw}%) vs losers ({rl}%) differs by {abs(delta):.0f} pp "
                f"— higher on {side}. Directional only (small n)."
            )
    return {
        "n_winners": len(winners),
        "n_losers": len(losers),
        "rate_winners_pct": rw,
        "rate_losers_pct": rl,
        "delta_pp_winners_minus_losers": delta,
        "discriminating": discriminating,
        "note": note,
    }


def validate_pre_entry_discrimination(ok: List[TradeAnalysis]) -> Dict[str, Any]:
    winners = [r for r in ok if (r.pnl_pts or 0) > 0]
    losers = [r for r in ok if (r.pnl_pts or 0) <= 0]
    combined = _flag_rate_block(
        winners, losers, lambda x: "PRE_ENTRY_WARNING" in x.flags
    )

    sub: Dict[str, Any] = {}
    for code, label in PRE_CODE_LABELS.items():
        block = _flag_rate_block(
            winners,
            losers,
            lambda x, c=code: c in (x.pre_entry.get("warning_codes") or []),
        )
        block["label"] = label
        sub[code] = block

    any_sub_disc = any(b.get("discriminating") for b in sub.values())
    if combined.get("discriminating") is False and not any_sub_disc:
        overall_note = (
            "PRE_ENTRY_WARNING (combined) does not discriminate winners from losers in this "
            "sample; no sub-flag clears the ~15 pp bar either. Do not use it as an entry filter yet."
        )
    elif combined.get("discriminating") is False and any_sub_disc:
        overall_note = (
            "Combined PRE_ENTRY_WARNING is not discriminating, but one or more sub-flags show a "
            "larger win/loss rate gap — see breakdown. Still directional only (n=15)."
        )
    elif combined.get("discriminating"):
        overall_note = combined.get("note") or ""
    else:
        overall_note = "Insufficient data to judge discrimination."

    return {
        "sample_caveat": (
            "n=15 trades. Win = P&L points > 0; loss/flat = P&L ≤ 0. Directional only."
        ),
        "combined": combined,
        "sub_flags": sub,
        "overall_note": overall_note,
    }


def summarize(results: List[TradeAnalysis]) -> Dict[str, Any]:
    ok = [r for r in results if not r.skip_reason]
    skipped = [r for r in results if r.skip_reason]

    by_grade: Dict[str, List[TradeAnalysis]] = {}
    by_sl: Dict[str, List[TradeAnalysis]] = {}
    for r in ok:
        by_grade.setdefault(r.trade["grade"], []).append(r)
        by_sl.setdefault(r.effective_sl_ref or "?", []).append(r)

    def _bucket(items: List[TradeAnalysis]) -> Dict[str, Any]:
        stable = [x for x in items if "DEGENERATE_RISK" not in x.flags]
        mae = [x.mae_mult for x in items if x.mae_mult is not None]
        mfe = [x.mfe_mult for x in items if x.mfe_mult is not None]
        mae_s = [x.mae_mult for x in stable if x.mae_mult is not None]
        mfe_s = [x.mfe_mult for x in stable if x.mfe_mult is not None]
        return {
            "n": len(items),
            "n_stable_risk": len(stable),
            "n_degenerate_risk": len(items) - len(stable),
            "round_trip": sum(1 for x in items if x.round_trip),
            "post_exit_reversion": sum(1 for x in items if x.post_exit_reversion),
            "avg_mae_mult": _avg(mae),
            "avg_mfe_mult": _avg(mfe),
            "median_mae_mult": _median(mae),
            "median_mfe_mult": _median(mfe),
            "avg_mae_mult_stable": _avg(mae_s),
            "avg_mfe_mult_stable": _avg(mfe_s),
            "avg_mae_pts": _avg([x.mae_pts for x in items if x.mae_pts is not None]),
            "avg_mfe_pts": _avg([x.mfe_pts for x in items if x.mfe_pts is not None]),
            "avg_mae_atr": _avg([x.mae_atr for x in items if x.mae_atr is not None]),
            "avg_mfe_atr": _avg([x.mfe_atr for x in items if x.mfe_atr is not None]),
            "single_candle_blowout": sum(
                1 for x in items if "SINGLE_CANDLE_MAE_BLOWOUT" in x.flags
            ),
            "single_candle_blowout_atr": sum(
                1 for x in items if x.single_candle_blowout_atr
            ),
            "pre_entry_warning": sum(1 for x in items if "PRE_ENTRY_WARNING" in x.flags),
        }

    return {
        "n_total": len(results),
        "n_analyzed": len(ok),
        "n_skipped": len(skipped),
        "skipped": [
            {"symbol": r.trade["symbol"], "date": r.trade["date"], "reason": r.skip_reason}
            for r in skipped
        ],
        "overall": _bucket(ok),
        "by_grade": {g: _bucket(v) for g, v in sorted(by_grade.items())},
        "by_sl_ref": {k: _bucket(v) for k, v in sorted(by_sl.items())},
        "denominator_check": validate_denominator_artifact(by_sl),
        "pre_entry_discrimination": validate_pre_entry_discrimination(ok),
        "sample_caveat": (
            "Sample is only 15 trades (13–17 Jul 2026). Findings are directional research signals, "
            "not statistically reliable conclusions."
        ),
    }


def render_html(results: List[TradeAnalysis], summary: Dict[str, Any], meta: Dict[str, Any]) -> str:
    def esc(s: Any) -> str:
        return (
            str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    rows_html = []
    for r in results:
        t = r.trade
        if r.skip_reason:
            rows_html.append(
                f"<tr class='skip'><td class='mono'>{esc(t['date'])}</td>"
                f"<td><b>{esc(t['symbol'])}</b></td><td>{esc(t['direction'])}</td>"
                f"<td colspan='10' class='empty'>SKIPPED — {esc(r.skip_reason)}</td></tr>"
            )
            continue
        flags = " ".join(
            f"<span class='pill pill-{esc(f.lower().split('_')[0])}'>{esc(f)}</span>"
            for f in r.flags
        ) or "<span class='meta'>—</span>"
        mae_note = "exit candle" if r.mae_on_exit_candle else (
            "earlier (recovered)" if r.mae_recovered else "earlier"
        )
        pre = r.pre_entry.get("summary") or "—"
        pre_cls = "tag-miss" if r.pre_entry.get("warning") else "meta"
        rev = (
            f"YES · {esc(r.post_exit_reversion_mult)}R"
            if r.post_exit_reversion
            else f"— <span class='meta'>({esc(r.post_exit_reversion_mult)}R)</span>"
        )
        rt = f"YES · {esc(r.round_trip_peak_mult)}R" if r.round_trip else "—"
        rows_html.append(
            f"""<tr>
            <td class="mono">{esc(t['date'])}<br><span class="meta">{esc(t['entry_time'])}→{esc(t['exit_time'])}</span></td>
            <td><b>{esc(t['symbol'])}</b><br><span class="meta">{esc(t['direction'])} · {esc(t['grade'])}</span></td>
            <td class="mono">{esc(t['entry_price'])} → {esc(t['exit_price'])}<br>
              <span class="meta">P&amp;L {esc(r.pnl_pts)} pts ({esc(r.pnl_r)}R)</span></td>
            <td class="mono">eff <b>{esc(r.effective_risk_pts)}</b> ({esc(r.effective_sl_ref)})<br>
              <span class="meta">EMA10 {esc(r.risk_ema10_pts)} · VWAP {esc(r.risk_vwap_pts)}</span><br>
              <span class="meta">ATR14 {esc(r.atr14_pts)} pts</span><br>
              <span class="meta">{esc(r.sl_choice_reason)}</span></td>
            <td class="mono"><b>{esc(r.mae_mult)}R</b> ({esc(r.mae_pts)} pts)<br>
              <span class="meta">{esc(r.mae_atr)}×ATR · {esc(mae_note)}</span></td>
            <td class="mono"><b>{esc(r.mfe_mult)}R</b> ({esc(r.mfe_pts)} pts)<br>
              <span class="meta">{esc(r.mfe_atr)}×ATR · @ +{esc(r.mfe_bars_after_entry)} bars</span></td>
            <td>{rt}</td>
            <td>{rev}</td>
            <td class="mono">{esc(r.exit_candle_range_mult)}R<br><span class="meta">{esc(r.exit_candle_range_pts)} pts</span></td>
            <td class="{pre_cls}">{esc(pre)}</td>
            <td>{flags}</td>
            <td class="meta">{esc((t.get('notes') or '')[:180])}</td>
            </tr>"""
        )

    o = summary.get("overall") or {}
    grade_cards = []
    for g, b in (summary.get("by_grade") or {}).items():
        grade_cards.append(
            f"""<div class="agg-card"><div class="k">Grade {esc(g)}</div>
            <div class="v">{esc(b['n'])}</div>
            <div class="s">MAE {esc(b.get('avg_mae_mult_stable') or b['avg_mae_mult'])}R ·
            MFE {esc(b.get('avg_mfe_mult_stable') or b['avg_mfe_mult'])}R (stable) ·
            RT {esc(b['round_trip'])} · Rev {esc(b['post_exit_reversion'])}</div></div>"""
        )
    sl_cards = []
    for k, b in (summary.get("by_sl_ref") or {}).items():
        sl_cards.append(
            f"""<div class="agg-card"><div class="k">SL ref {esc(k)}</div>
            <div class="v">{esc(b['n'])}</div>
            <div class="s">MAE {esc(b.get('avg_mae_mult_stable') or b['avg_mae_mult'])}R ·
            MFE {esc(b.get('avg_mfe_mult_stable') or b['avg_mfe_mult'])}R (stable) ·
            RT {esc(b['round_trip'])} · Rev {esc(b['post_exit_reversion'])}</div></div>"""
        )

    skip_html = ""
    if summary.get("skipped"):
        items = "".join(
            f"<li><b>{esc(s['symbol'])}</b> {esc(s['date'])} — {esc(s['reason'])}</li>"
            for s in summary["skipped"]
        )
        skip_html = f'<div class="warn"><b>Skipped rows</b><ul>{items}</ul></div>'

    dc = summary.get("denominator_check") or {}
    ema_m = dc.get("ema10") or {}
    vwap_m = dc.get("vwap") or {}
    gaps = dc.get("gaps") or {}
    vlabel = dc.get("verdict_label") or ""
    vcls = {
        "artifact": "verdict-artifact",
        "persists": "verdict-persists",
        "no_r_split": "verdict-mixed",
        "mixed": "verdict-mixed",
    }.get(vlabel, "verdict-mixed")

    def _mrow(label: str, ek: str, fmt: str = "") -> str:
        ev, vv = ema_m.get(ek), vwap_m.get(ek)
        return (
            f"<tr><td>{esc(label)}</td>"
            f"<td class='mono'>{esc(ev)}</td><td class='mono'>{esc(vv)}</td></tr>"
        )

    denom_html = f"""
  <details class="panel" open>
    <summary>Validation 1 — Is the VWAP-worse split a denominator artifact?</summary>
    <div class="panel-body">
      <p class="meta">{esc(dc.get('sample_caveat'))}</p>
      <div class="{esc(vcls)}"><b>Verdict:</b> {esc(dc.get('verdict'))}</div>
      <p class="meta" style="margin-top:8px">Relative gaps (VWAP vs EMA10 group; positive = VWAP higher):
        avg risk {esc(gaps.get('vwap_vs_ema_avg_risk_pts'))} ·
        MAE R {esc(gaps.get('vwap_vs_ema_avg_mae_r_stable'))} ·
        avg MAE pts {esc(gaps.get('vwap_vs_ema_avg_mae_pts'))} ·
        median MAE pts {esc(gaps.get('vwap_vs_ema_median_mae_pts'))} ·
        MAE/ATR {esc(gaps.get('vwap_vs_ema_avg_mae_atr'))}</p>
      <table>
        <thead><tr><th>Metric</th><th>EMA10-effective</th><th>VWAP-effective</th></tr></thead>
        <tbody>
          {_mrow('n', 'n')}
          {_mrow('Avg / median planned risk (pts)', 'avg_effective_risk_pts')}
          <tr><td class="meta">median risk</td>
            <td class="mono">{esc(ema_m.get('median_effective_risk_pts'))}</td>
            <td class="mono">{esc(vwap_m.get('median_effective_risk_pts'))}</td></tr>
          {_mrow('Avg MAE (pts)', 'avg_mae_pts')}
          {_mrow('Median MAE (pts)', 'median_mae_pts')}
          {_mrow('Avg MFE (pts)', 'avg_mfe_pts')}
          {_mrow('Avg exit-candle range (pts)', 'avg_exit_range_pts')}
          {_mrow('Avg ATR(14) at entry (pts)', 'avg_atr14_pts')}
          {_mrow('Avg MAE / ATR', 'avg_mae_atr')}
          {_mrow('Median MAE / ATR', 'median_mae_atr')}
          {_mrow('Avg MFE / ATR', 'avg_mfe_atr')}
          {_mrow('Avg MAE R (stable)', 'avg_mae_mult_stable')}
          {_mrow('Blowouts MAE≥2R on exit bar', 'single_candle_blowout_r')}
          {_mrow('Blowouts MAE≥2×ATR on exit bar', 'single_candle_blowout_atr')}
          {_mrow('Post-exit reversion count', 'post_exit_reversion')}
          {_mrow('Round-trip count', 'round_trip')}
        </tbody>
      </table>
    </div>
  </details>
"""

    pd = summary.get("pre_entry_discrimination") or {}
    comb = pd.get("combined") or {}
    sub_rows = []
    for code, block in (pd.get("sub_flags") or {}).items():
        disc = block.get("discriminating")
        disc_s = "yes" if disc is True else ("no" if disc is False else "—")
        sub_rows.append(
            f"<tr><td>{esc(block.get('label') or code)}</td>"
            f"<td class='mono'>{esc(block.get('rate_winners_pct'))}%</td>"
            f"<td class='mono'>{esc(block.get('rate_losers_pct'))}%</td>"
            f"<td class='mono'>{esc(block.get('delta_pp_winners_minus_losers'))} pp</td>"
            f"<td>{esc(disc_s)}</td></tr>"
        )
    pre_html = f"""
  <details class="panel" open>
    <summary>Validation 2 — Does PRE_ENTRY_WARNING discriminate winners from losers?</summary>
    <div class="panel-body">
      <p class="meta">{esc(pd.get('sample_caveat'))}</p>
      <div class="verdict-mixed"><b>Combined flag:</b> {esc(pd.get('overall_note') or comb.get('note'))}</div>
      <div class="agg" style="margin-top:10px">
        <div class="agg-card"><div class="k">Winners (P&amp;L&gt;0)</div>
          <div class="v">{esc(comb.get('n_winners'))}</div>
          <div class="s">flag rate {esc(comb.get('rate_winners_pct'))}%</div></div>
        <div class="agg-card"><div class="k">Losers / flat (P&amp;L≤0)</div>
          <div class="v">{esc(comb.get('n_losers'))}</div>
          <div class="s">flag rate {esc(comb.get('rate_losers_pct'))}%</div></div>
        <div class="agg-card"><div class="k">Rate gap</div>
          <div class="v">{esc(comb.get('delta_pp_winners_minus_losers'))} pp</div>
          <div class="s">winners − losers</div></div>
      </div>
      <h2 style="margin-top:12px">Sub-flag rates</h2>
      <table>
        <thead><tr>
          <th>Sub-flag</th><th>Winners</th><th>Losers/flat</th><th>Δ pp</th><th>Discriminates?</th>
        </tr></thead>
        <tbody>{''.join(sub_rows) or '<tr><td colspan="5" class="empty">—</td></tr>'}</tbody>
      </table>
    </div>
  </details>
"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Trade MAE/MFE Excursion — 13–17 Jul 2026</title>
  <style>
    :root {{
      --bg: #0f1419; --panel: #1a222c; --panel-2: #232d3a; --border: #334155;
      --text: #e8eef4; --muted: #94a3b8; --accent: #0d9488;
      --font: "IBM Plex Sans", "Segoe UI", system-ui, sans-serif;
      --mono: "IBM Plex Mono", ui-monospace, Menlo, monospace;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0; font-family: var(--font); color: var(--text); font-size: 13px; line-height: 1.4;
      background: radial-gradient(1200px 600px at 10% -10%, #1a2a32 0%, var(--bg) 55%);
      min-height: 100vh;
    }}
    .wrap {{ max-width: 1400px; margin: 0 auto; padding: 16px 18px 48px; }}
    header {{
      display: flex; flex-wrap: wrap; justify-content: space-between; gap: 12px;
      border-bottom: 1px solid var(--border); padding-bottom: 12px; margin-bottom: 14px;
    }}
    h1 {{ margin: 0; font-size: 1.25rem; font-weight: 700; }}
    .sub {{ margin: 4px 0 0; color: var(--muted); font-size: 12px; max-width: 70ch; }}
    .agg {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 8px; margin-bottom: 14px; }}
    .agg-card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; }}
    .agg-card .k {{ color: var(--muted); font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.04em; }}
    .agg-card .v {{ font-size: 1.35rem; font-weight: 700; font-variant-numeric: tabular-nums; margin-top: 2px; }}
    .agg-card .s {{ font-size: 11px; color: var(--muted); margin-top: 2px; }}
    .tag-miss {{ color: #fdba74; }}
    .meta {{ color: var(--muted); font-size: 12px; }}
    .mono {{ font-family: var(--mono); font-size: 12px; }}
    .warn {{
      background: rgba(217, 119, 6, 0.15); border: 1px solid rgba(217, 119, 6, 0.4);
      color: #fde68a; padding: 8px 10px; border-radius: 6px; margin-bottom: 10px;
    }}
    details.panel {{
      background: var(--panel); border: 1px solid var(--border); border-radius: 8px; margin-bottom: 12px;
    }}
    details.panel > summary {{
      cursor: pointer; padding: 10px 12px; font-weight: 700; list-style: none;
    }}
    details.panel[open] > summary {{ border-bottom: 1px solid var(--border); }}
    .panel-body {{ padding: 12px; overflow-x: auto; }}
    table {{ width: 100%; border-collapse: collapse; font-variant-numeric: tabular-nums; }}
    th, td {{ text-align: left; padding: 6px 8px; border-bottom: 1px solid var(--border); vertical-align: top; }}
    th {{ color: var(--muted); font-size: 10px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.03em; }}
    tr.skip td {{ opacity: 0.7; }}
    .pill {{
      display: inline-block; padding: 1px 7px; border-radius: 999px; font-size: 10px;
      font-weight: 700; letter-spacing: 0.03em; margin: 1px 2px 1px 0;
      background: rgba(13, 148, 136, 0.25); color: #5eead4;
    }}
    .pill-round {{ background: rgba(234, 88, 12, 0.22); color: #fdba74; }}
    .pill-post {{ background: rgba(192, 38, 211, 0.22); color: #f0abfc; }}
    .pill-single, .pill-exit {{ background: rgba(220, 38, 38, 0.22); color: #fecaca; }}
    .pill-pre {{ background: rgba(202, 138, 4, 0.22); color: #fcd34d; }}
    .pill-degenerate {{ background: rgba(100, 116, 139, 0.35); color: #cbd5e1; }}
    .empty {{ color: var(--muted); font-style: italic; }}
    h2 {{ font-size: 0.95rem; margin: 18px 0 8px; }}
    .verdict-artifact {{
      background: rgba(13, 148, 136, 0.15); border: 1px solid rgba(13, 148, 136, 0.45);
      color: #5eead4; padding: 10px 12px; border-radius: 6px; margin: 8px 0 12px;
    }}
    .verdict-persists {{
      background: rgba(220, 38, 38, 0.12); border: 1px solid rgba(220, 38, 38, 0.4);
      color: #fecaca; padding: 10px 12px; border-radius: 6px; margin: 8px 0 12px;
    }}
    .verdict-mixed {{
      background: rgba(217, 119, 6, 0.15); border: 1px solid rgba(217, 119, 6, 0.4);
      color: #fde68a; padding: 10px 12px; border-radius: 6px; margin: 8px 0 12px;
    }}
  </style>
</head>
<body>
<div class="wrap">
  <header>
    <div>
      <h1>Trade MAE / MFE Excursion</h1>
      <p class="sub">13–17 Jul 2026 actual trades · 10m Upstox candles · effective planned risk =
      nearer of EMA10 vs VWAP (Notes override when explicit). Research only — no live gating.</p>
    </div>
    <div class="meta">Generated {esc(meta.get('generated_at'))}<br>Source {esc(meta.get('xlsx'))}</div>
  </header>

  {skip_html}
  <div class="warn"><b>Sample caveat:</b> {esc(summary.get('sample_caveat'))}
  Headline MAE/MFE R averages exclude <b>DEGENERATE_RISK</b> rows
  (effective SL gap &lt; max(0.5 pts, 0.05% of entry)). Per-trade R still shown.</div>

  <div class="agg">
    <div class="agg-card"><div class="k">Analyzed</div><div class="v">{esc(o.get('n'))}</div>
      <div class="s">of {esc(summary.get('n_total'))} rows · stable {esc(o.get('n_stable_risk'))}</div></div>
    <div class="agg-card"><div class="k">Avg MAE (stable)</div><div class="v">{esc(o.get('avg_mae_mult_stable'))}R</div>
      <div class="s">median {esc(o.get('median_mae_mult'))}R · raw avg {esc(o.get('avg_mae_mult'))}R</div></div>
    <div class="agg-card"><div class="k">Avg MFE (stable)</div><div class="v">{esc(o.get('avg_mfe_mult_stable'))}R</div>
      <div class="s">median {esc(o.get('median_mfe_mult'))}R · raw avg {esc(o.get('avg_mfe_mult'))}R</div></div>
    <div class="agg-card"><div class="k">Round-trip</div><div class="v">{esc(o.get('round_trip'))}</div>
      <div class="s">≥1.5R then BE/loss</div></div>
    <div class="agg-card"><div class="k">Post-exit rev.</div><div class="v">{esc(o.get('post_exit_reversion'))}</div>
      <div class="s">≥1R resume after exit</div></div>
    <div class="agg-card"><div class="k">1-candle blowout</div><div class="v">{esc(o.get('single_candle_blowout'))}</div>
      <div class="s">MAE≥2R on exit bar</div></div>
    <div class="agg-card"><div class="k">Pre-entry warn</div><div class="v">{esc(o.get('pre_entry_warning'))}</div>
      <div class="s">structure red flags</div></div>
  </div>

  <h2>By confidence grade</h2>
  <div class="agg">{''.join(grade_cards) or '<div class="meta">—</div>'}</div>

  <h2>By effective SL reference</h2>
  <div class="agg">{''.join(sl_cards) or '<div class="meta">—</div>'}</div>

  {denom_html}
  {pre_html}

  <details class="panel" open>
    <summary>Per-trade excursion table</summary>
    <div class="panel-body">
      <table>
        <thead>
          <tr>
            <th>Date / time</th>
            <th>Symbol</th>
            <th>Entry → Exit</th>
            <th>Planned risk</th>
            <th>MAE</th>
            <th>MFE</th>
            <th>Round-trip</th>
            <th>Post-exit rev</th>
            <th>Exit bar range</th>
            <th>Pre-entry</th>
            <th>Flags</th>
            <th>Notes</th>
          </tr>
        </thead>
        <tbody>
          {''.join(rows_html)}
        </tbody>
      </table>
    </div>
  </details>

  <p class="meta">Effective risk = min(|entry−EMA10|, |entry−VWAP|) unless Notes force a reference.
  MAE/MFE from 10m OHLC between entry and exit bars. Post-exit window = next {POST_BARS} ten-minute bars.
  Pre-entry = prior {PRE_BARS} bars (EMA10 vs VWAP + volume/body deceleration heuristics).
  ATR(14) = Wilder ATR on multi-day 10m bars at the entry bar. Sample n=15 — directional only.</p>
</div>
</body>
</html>
"""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"Excel not found: {args.xlsx}", file=sys.stderr)
        return 1

    trades = load_trades(args.xlsx)
    print(f"Loaded {len(trades)} trades from {args.xlsx}")

    from backend.database import SessionLocal

    db = SessionLocal()
    results: List[TradeAnalysis] = []
    candle_cache: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    try:
        for tr in trades:
            key = (tr.symbol, tr.session_date.isoformat())
            print(
                f"  {tr.session_date} {tr.symbol} {tr.direction} {tr.entry_time}…",
                flush=True,
            )
            ikey = resolve_instrument_key(db, tr.symbol)
            if not ikey:
                results.append(
                    TradeAnalysis(
                        trade=_trade_dict(tr),
                        skip_reason="no futures instrument_key in arbitrage_master",
                    )
                )
                continue
            if key not in candle_cache:
                try:
                    c5 = fetch_5m_candles(ikey, tr.session_date)
                    candle_cache[key] = build_session_10m(c5, tr.session_date)
                    print(f"    candles: {len(c5)} 5m → {len(candle_cache[key])} 10m")
                except Exception as exc:
                    print(f"    candle fetch failed: {exc}")
                    candle_cache[key] = []
            results.append(analyze_one(tr, candle_cache[key], ikey))
    finally:
        db.close()

    summary = summarize(results)
    meta = {
        "generated_at": datetime.now(IST).isoformat(),
        "xlsx": str(args.xlsx),
        "pre_bars": PRE_BARS,
        "post_bars": POST_BARS,
    }
    args.out_dir.mkdir(parents=True, exist_ok=True)
    stem = "TRADE_MAE_MFE_13JUL_17JUL2026"
    html_path = args.out_dir / f"{stem}.html"
    json_path = args.out_dir / f"{stem}.json"
    payload = {
        "meta": meta,
        "summary": summary,
        "trades": [asdict(r) for r in results],
    }
    json_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    html_path.write_text(render_html(results, summary, meta), encoding="utf-8")
    public_html = _ROOT / "frontend/public/trade-mae-mfe.html"
    public_html.write_text(html_path.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"Wrote {html_path}")
    print(f"Wrote {json_path}")
    print(f"Wrote {public_html}")
    print(json.dumps(summary.get("overall"), indent=2))
    dc = summary.get("denominator_check") or {}
    print("DENOMINATOR:", dc.get("verdict_label"), (dc.get("verdict") or "")[:200])
    pe = summary.get("pre_entry_discrimination") or {}
    print("PRE_ENTRY:", (pe.get("overall_note") or "")[:200])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
